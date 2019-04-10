# -*- coding: utf-8 -*-
from django.shortcuts import render, get_object_or_404, redirect
from django import forms
from django.db import connections
from django.contrib import messages
from django.contrib.auth.models import *
from django.http import HttpResponse, Http404, HttpResponseRedirect, JsonResponse, HttpResponseServerError, StreamingHttpResponse
from django.template import RequestContext, loader
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.urlresolvers import reverse, reverse_lazy
from django.core.exceptions import ObjectDoesNotExist
from django.forms import formset_factory, BaseFormSet
from django.forms.models import  modelformset_factory, inlineformset_factory, model_to_dict
from django.views import generic
from django.db.models.signals import pre_save
from django.db.models import Sum, Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.core.serializers.json import DjangoJSONEncoder
from django.template.defaultfilters import filesizeformat
from django.utils import timezone
from datetime import datetime  
from django.utils.timezone import now, utc 
from django.core.mail import send_mail
from django.core import serializers
from django.core.files import File
from django.db import IntegrityError
from .forms import *
from openpyxl import load_workbook
import os
import win32com.client
import math
from win32com.client import constants
from win32com.client import gencache
import win32gui
import win32process
import pythoncom
import PIL.ImageGrab
from PIL import Image
import pywinauto
import re
import win32con
import time

# Create your views here.


def home(request):
	# load homepage html template
	template = loader.get_template('FreedAM_app/templates/FreedAM_app/homepage.html')

	# contact_form = ContactForm(request.POST or None)
	# if contact_form.is_valid():
	# 	contact_form.save()
	context = {
	  # 'contact_form': contact_form,
	}
	return HttpResponse(template.render(context, request))	

def projects(request):
	# load projects page 
	template = loader.get_template('FreedAM_app/templates/FreedAM_app/projects.html')
	context = {
	   # 'latest_question_list': latest_question_list,
	}
	return HttpResponse(template.render(context, request))	

# frame dimension calculator page
def calculator_home(request):
	# load calculator html page 
	template = loader.get_template('FreedAM_app/templates/FreedAM_app/calculator_home.html')
	# load input form 
	frame_input_form = FrameDimensionsForm(request.POST or None)
	#initialize pythoncom to allow connection to Inventor
	pythoncom.CoInitialize()	
	# Open Inventor
	invApp = win32com.client.Dispatch("Inventor.Application")

	# Make inventor visible
	invApp.Visible = True

	# Operate Inventor in the background
	invApp.SilentOperation = True

	# Set location of assembly
	Assembly_name = 'C:/Users/asidawi/Desktop/CAD_testing/Frame_assembly - Working_on_constraints.iam'

	# if the form is valid i.e. has no clear input errors
	if frame_input_form.is_valid():

		# Open the model
		oDoc = invApp.Documents.Open(Assembly_name)		

		# get latest object created id - THIS IS NOT ROBUST!!!!!!!!!!!!!!!!!!!!!!!!
		object_created_id = FrameDimensions.objects.order_by('-id')[0].id
		
		# if angles are greater than 105 or less than 85 degrees raise error as this goes against anthropometric data
		if frame_input_form.cleaned_data['angle_lower_leg_upper_leg'] > 105 or frame_input_form.cleaned_data['angle_lower_leg_upper_leg'] < 85:
			form_errors = "Please ensure that angles are greater than 85 or less than 105 degrees"
			context = {'form_errors': form_errors, 'frame_input_form': frame_input_form}
			return HttpResponse(template.render(context, request))
		# seat depth range is between 400 and 500mm due to availability of external components. Theoretically could be any value but cost of seat would increase	
		if frame_input_form.cleaned_data['seat_depth'] > 500 or frame_input_form.cleaned_data['seat_depth'] < 400:
			form_errors = "Please ensure that seat depth is between 400 and 500mm"
			context = {'form_errors': form_errors, 'frame_input_form': frame_input_form}
			return HttpResponse(template.render(context, request))	
		# seat width between 400 and 600mm due to same reason as above and to prevent footrest parts touching
		if frame_input_form.cleaned_data['seat_width'] > 600 or frame_input_form.cleaned_data['seat_width'] < 400:
			form_errors = "Please ensure that seat width is between 400 and 600mm"
			context = {'form_errors': form_errors, 'frame_input_form': frame_input_form}
			return HttpResponse(template.render(context, request))	
		# seat height between 400 and 500mm
		if frame_input_form.cleaned_data['seat_height'] > 500 or frame_input_form.cleaned_data['seat_depth'] < 400:
			form_errors = "Please ensure that seat height is between 400 and 500mm"
			context = {'form_errors': form_errors, 'frame_input_form': frame_input_form}
			return HttpResponse(template.render(context, request))	
		# raise error if shoulder height less than seat height or shoulder height greater than 1500 
		if frame_input_form.cleaned_data['shoulder_height'] < frame_input_form.cleaned_data['seat_height'] or frame_input_form.cleaned_data['shoulder_height'] >1500:
			form_errors = "Please ensure that shoulder height is greater than seat height, and less than 1500mm"
			context = {'form_errors': form_errors, 'frame_input_form': frame_input_form}
			return HttpResponse(template.render(context, request))													
		
		# load spreadsheet with parameters stored
		os.chdir("C:/Users/asidawi/Desktop/CAD_testing")	
		wb = load_workbook('Parameters.xlsx')

		# load Sheet1
		ws = wb['Sheet1']
		
		# assign spreadsheet values according to input data
		ws['B2'] = frame_input_form.cleaned_data['angle_lower_leg_upper_leg']
		ws['B3'] = frame_input_form.cleaned_data['seating_angle']
		ws['B4'] = frame_input_form.cleaned_data['backrest_angle']

		# use standard upholstery sizes and compare backrest width and length to that (6 standard sizes) 400x400, 400x450, 400x500, 450x550, 450x600

		# backrest tube length should be equal to shoulder height - seat height
		backrest_tube_length = frame_input_form.cleaned_data['shoulder_height'] - frame_input_form.cleaned_data['seat_height']
		
		# add 25mm buffer to backrest tube length due to insert into joints
		if backrest_tube_length <= 400:
			ws['B5'] = 400 + 25
		elif 401 <=  backrest_tube_length <= 450:
			ws['B5'] = 450 + 25
		elif 451 <=  backrest_tube_length <= 500:
			ws['B5'] = 500 + 25	
		elif 501 <=  backrest_tube_length <= 550:
			ws['B5'] = 550 + 25					
		else:
			ws['B5'] = 600 + 25	
			
		# calculate whether vertical tubes need to be extended 
		seat_height = frame_input_form.cleaned_data['seat_height']
		seat_depth = frame_input_form.cleaned_data['seat_depth']
		vertical_tube_length_rear = seat_height - 300
		vertical_tube_length_front = seat_height - 300
		seat_angle = frame_input_form.cleaned_data['seating_angle']
		distance_between_rear_joints_and_front_joints = frame_input_form.cleaned_data['seat_depth'] + 0.1*frame_input_form.cleaned_data['seat_depth'] - 40 + 25	
		
		# conditions for changing vertical tube angle and calculating the new length of vertical tube and 
		if seat_angle > 90:
			angle_difference = seat_angle - 90
			additional_vertical_tube_length = math.tan(math.radians(angle_difference))*distance_between_rear_joints_and_front_joints
			vertical_tube_length_rear = vertical_tube_length_rear + additional_vertical_tube_length
			new_seat_tube_length = math.sqrt(additional_vertical_tube_length**2 + seat_depth**2)
		elif seat_angle < 90:
			angle_difference = 90 - seat_angle
			additional_vertical_tube_length = math.tan(math.radians(angle_difference))*distance_between_rear_joints_and_front_joints
			vertical_tube_length_front = vertical_tube_length_front + additional_vertical_tube_length
			new_seat_tube_length = math.sqrt(additional_vertical_tube_length**2 + seat_depth**2)			
		else:
			pass

		# setting remainder spreadsheet values from form inputs
		ws['B7'] = vertical_tube_length_front
		ws['B8'] = vertical_tube_length_rear

		# set seat depth
		# use standard upholstery sizes and compare seat width and length to that (6 standard sizes) 400x400, 400x450, 400x500, 450x550, 450x600
		# length of horizontal tubes are seat depth + 50 to account for insertion into joints
		if frame_input_form.cleaned_data['seat_depth'] <= 400:
			ws['B9'] = 400 + 50	 
			ws['B6'] = 400 + 50			
		elif 401 <=  frame_input_form.cleaned_data['seat_depth'] <= 450:
			ws['B9'] = 450 + 50	
			ws['B6'] = 450 + 50						
		elif 451 <=  frame_input_form.cleaned_data['seat_depth'] <= 500:
			ws['B9'] = 500 + 50		
			ws['B6'] = 500 + 50							
		elif 501 <=  frame_input_form.cleaned_data['seat_depth'] <= 550:
			ws['B9'] = 550 + 50	
			ws['B6'] = 550 + 50											
		else:
			ws['B9'] = 600 + 50		
			ws['B6'] = 600 + 50			

		# divide seat width by 2 to get cross member lengths
		ws['B10'] = frame_input_form.cleaned_data['seat_width']/2 - 25
		ws['B11'] = frame_input_form.cleaned_data['seat_width']/2 - 25

		# cross_member_length = math.sqrt((seat_height - 300)**2 + seat_width**2)
		# ws['B10'] = cross_member_length
		# ws['B11'] = math.degrees(math.atan(seat_width/(seat_height-250)))

		# set seat width values
		if frame_input_form.cleaned_data['seat_width'] <= 400:
			ws['B14'] = 400
		elif 401 <=  frame_input_form.cleaned_data['seat_width'] <= 450:
			ws['B14'] = 450
		else:
			ws['B14'] = 500		

		# save parameters spreadsheet
		wb.save('Parameters.xlsx')	

		# iLogic GUID
		iLogicAddinGuid = "{3BDD8D79-2179-4B11-8A5A-257B1C0263AC}"

		# load iLogic add-in
		iLogic_addin = invApp.ApplicationAddIns.ItemById(iLogicAddinGuid)

		iLogicAutomation = iLogic_addin.Automation()
		# Update assembly
		oDoc.Update()

		# wait 5 seconds
		time.sleep(5)

		# Update assembly again
		oDoc.Update()

		# wait 10 seconds
		time.sleep(10)

		# update again to ensure updates have been applied
		oDoc.Update()

		# Save assembly
		oDoc.Save()

		#rules = iLogicAutomation.get_Rules(oDoc)

		#backrest_joint = 'C:/Users/asidawi/Desktop/CAD_testing/Backrest_joint.ipt'

		#oDoc_bj = invApp.Documents.Open(backrest_joint)		

		# Save joints as STL (needs testing)
		#oDoc_bj.SaveAs("backrest_joint_"+str(object_created_id)+".stl", True)

		# change directory to FreedAM image folder
		os.chdir("C:/Users/asidawi/Documents/FreedAM/FreedAM/FreedAM_app/static/images")	

		# function to fetch window handle number
		def enumHandler(hwnd, lParam):
			if win32gui.IsWindowVisible(hwnd):
				# if inventor is in the title of the window
				if 'Inventor' in win32gui.GetWindowText(hwnd):
					# minimize window 
					win32gui.ShowWindow(hwnd, win32con.SW_MINIMIZE)		
					# maximize window to bring it to the front		
					win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)

		win32gui.EnumWindows(enumHandler, None)

		# set camera in inventor to right hand side view
		Right_view = invApp.CommandManager.ControlDefinitions.Item("AppRightViewCmd").Execute()
		# take screenshot and crop to wheelchair size
		im = PIL.ImageGrab.grab()   
		im.crop((500, 150, 1500, 980)).save("FreedAM_"+str(object_created_id)+"_right.jpg")

		# set camera in inventor to front view
		Front_view = invApp.CommandManager.ControlDefinitions.Item("AppFrontViewCmd").Execute()
		# take screenshot and crop to wheelchair size
		im = PIL.ImageGrab.grab()     
		im.crop((500, 150, 1500, 980)).save("FreedAM_"+str(object_created_id)+"_front.jpg")

		# minimize Inventor window
		win32gui.ShowWindow(7346022, win32con.SW_MINIMIZE)				

		# save frame input form data to database
		frame_input_form.save()

		# create an accessories object for the next page linked to the new frame
		new_accessories_object = Accessories(pub_date = datetime.datetime.now(), linked_frame = FrameDimensions.objects.get(id =object_created_id))
		# save accessories object
		new_accessories_object.save()

		# redirect to frame_preview page
		return redirect("FreedAM_app:frame_preview", object_created_id)
	else:
		# if form is not valid return errors to webpage
		form_errors = frame_input_form.errors
		print "form invalid"
		context = {'form_errors': form_errors, 'frame_input_form': frame_input_form}
		return HttpResponse(template.render(context, request))		

	context = {
	   'frame_input_form': frame_input_form,
	}
	return HttpResponse(template.render(context, request))	

# frame preview and option selection page
def frame_preview(request, id=None):

	# load frame preview template file
	template = loader.get_template('FreedAM_app/templates/FreedAM_app/frame_preview.html')

	# load relative FrameDimensions object
	measurements = FrameDimensions.objects.get(id=id)

	# load relevant wheelchair images
	front_image_location = '/static/images/FreedAM_'+str(id)+'_front.jpg'
	right_image_location = '/static/images/FreedAM_'+str(id)+'_right.jpg'

	# load accessories form
	accessories_form = AccessoriesForm(request.POST or None, instance = measurements)

	# if form is valid save it
	if accessories_form.is_valid():
		accessories_form.linked_frame = measurements
		accessories_form.save()
		return redirect("FreedAM_app:homepage")

	# load preview image
	context = {
	   'measurements': measurements, 'front_image_location': front_image_location, 'right_image_location': right_image_location, 'accessories_form': accessories_form
	}
	return HttpResponse(template.render(context, request))	
	#return render(request, "buildpage/add_note.html", context)

# FreedAM project description page
def FreedAM_project(request):
	template = loader.get_template('FreedAM_app/templates/FreedAM_app/FreedAM_project.html')

	context = {
	  # 'contact_form': contact_form,
	}
	return HttpResponse(template.render(context, request))	